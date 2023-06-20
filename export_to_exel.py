from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, numbers
from openpyxl.styles.builtins import styles
from openpyxl.worksheet.table import Table, TableStyleInfo
from data import alldata, sum, earn_sum


def export():
    all_data = alldata()
    summa = sum()
    earn = earn_sum()
    data = []

    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    for i in all_data:
        data.append(list(i))

    # print(data)
    first_date = data[0][0]
    last_date = data[-1][0]

    data_len = len(data) + 1

    ws.append(["Дата", "Наименование", "Техника", "Доп. инф.", "Цена", "Приход"])

    column_dimension = ws.column_dimensions["A"]
    column_dimension.number_format = numbers.FORMAT_DATE_DATETIME


    for row in data:
        ws.append(row)

    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    title_cells = ["A1", "B1", "C1", "D1", "E1", "F1"]
    for i in title_cells:
        cell1 = ws[i]
        cell1.style = styles["Good"]
        cell1.alignment = Alignment(horizontal='center', vertical='center')



    rows = 'A1:F' + str(data_len+1)
    range = ws[rows]
    for row in range:
        for cell in row:
            cell.border = border

    merge = 'B' + str(data_len + 1) + ":" + "C" + str(data_len + 1)
    ws.merge_cells(merge)
    ws.merge_cells('B' + str(data_len + 1) + ":" + "D" + str(data_len + 1))
    ws["B" + str(data_len + 1)] = "Итого"
    ws["B" + str(data_len + 1)].alignment = Alignment(horizontal='center', vertical='center')

    ws["E" + str(data_len + 1)] = int(summa[0][0])
    ws["F" + str(data_len + 1)] = int(earn[0][0])
    name = "Отчёт" + "_" + str(first_date) + "_" + str(last_date) + ".xlsx"
    wb.save(name)
    return name
